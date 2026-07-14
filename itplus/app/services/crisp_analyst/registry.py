"""Register tabular documents into DuckDB on indexing (CRISP-DM: Data Understanding)."""

from __future__ import annotations

import logging
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

import duckdb
from sqlalchemy.orm import Session

from itplus.app.core.config import get_settings
from itplus.app.models.document import Document
from itplus.app.services.crisp_analyst.models import DatasetProfile

logger = logging.getLogger(__name__)

_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("date", "fecha", "mes", "periodo", "datetime"),
    "quantity": ("quantity", "cantidad", "qty", "unidades", "units"),
    "price": ("price", "precio", "monto", "amount", "unit_price"),
    "category": ("category", "categoria", "familia", "tipo", "segment"),
    "city": ("city", "ciudad", "region", "zona", "location"),
    "product": ("product", "producto", "item", "sku", "nombre"),
    "order_id": ("order_id", "orderid", "id_pedido", "pedido"),
}


def _analyst_dir() -> Path:
    base = Path(get_settings().upload_dir) / "analyst"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _duckdb_path(document_id: uuid.UUID) -> Path:
    return _analyst_dir() / f"{document_id}.duckdb"


def _norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_")


def _detect_column(columns: list[str], role: str) -> str | None:
    aliases = _COLUMN_ALIASES.get(role, ())
    normalized = {_norm(c): c for c in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    for col in columns:
        n = _norm(col)
        if any(alias in n for alias in aliases):
            return col
    return None


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def register_document_dataset(db: Session, document: Document) -> DatasetProfile | None:
    """Load CSV/Excel into DuckDB and persist analyst_profile on the document."""
    path = Path(document.storage_path)
    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xlsm"}:
        return None

    duck_path = _duckdb_path(document.id)
    con = duckdb.connect(str(duck_path))

    try:
        if suffix == ".csv":
            con.execute(
                f"CREATE OR REPLACE TABLE raw_data AS SELECT * FROM read_csv_auto({_quote_ident(str(path))})"
            )
        else:
            csv_temp = _analyst_dir() / f"{document.id}_import.csv"
            _xlsx_to_csv(path, csv_temp)
            con.execute(
                f"CREATE OR REPLACE TABLE raw_data AS SELECT * FROM read_csv_auto({_quote_ident(str(csv_temp))})"
            )

        columns = [row[0] for row in con.execute("DESCRIBE raw_data").fetchall()]
        if not columns:
            return None

        date_col = _detect_column(columns, "date")
        qty_col = _detect_column(columns, "quantity")
        price_col = _detect_column(columns, "price")
        category_col = _detect_column(columns, "category")
        city_col = _detect_column(columns, "city")
        product_col = _detect_column(columns, "product")

        revenue_parts: list[str] = []
        if qty_col and price_col:
            revenue_parts.append(
                f"TRY_CAST({_quote_ident(qty_col)} AS DOUBLE) * TRY_CAST({_quote_ident(price_col)} AS DOUBLE)"
            )
        if price_col:
            revenue_parts.append(f"TRY_CAST({_quote_ident(price_col)} AS DOUBLE)")
        revenue_expr = revenue_parts[0] if revenue_parts else "0"

        dt_expr = f"TRY_CAST({_quote_ident(date_col)} AS DATE)" if date_col else "NULL"
        month_expr = f"strftime({dt_expr}, '%Y-%m')" if date_col else "NULL"

        con.execute(
            f"""
            CREATE OR REPLACE VIEW data_clean AS
            SELECT
                *,
                {revenue_expr} AS revenue,
                {dt_expr} AS dt,
                {month_expr} AS month_key
            FROM raw_data
            """
        )

        row_count = int(con.execute("SELECT COUNT(*) FROM data_clean").fetchone()[0])
        date_min = date_max = None
        if date_col:
            bounds = con.execute(
                "SELECT MIN(dt), MAX(dt) FROM data_clean WHERE dt IS NOT NULL"
            ).fetchone()
            if bounds and bounds[0]:
                date_min = str(bounds[0])
                date_max = str(bounds[1])

        profile = DatasetProfile(
            document_id=document.id,
            document_name=document.filename,
            row_count=row_count,
            columns=columns,
            date_column=date_col,
            quantity_column=qty_col,
            price_column=price_col,
            category_column=category_col,
            city_column=city_col,
            product_column=product_col,
            date_min=date_min,
            date_max=date_max,
            revenue_expression="revenue",
        )

        document.analyst_profile = {
            "registered_at": datetime.now(timezone.utc).isoformat(),
            "row_count": row_count,
            "columns": columns,
            "date_column": date_col,
            "quantity_column": qty_col,
            "price_column": price_col,
            "category_column": category_col,
            "city_column": city_col,
            "product_column": product_col,
            "date_min": date_min,
            "date_max": date_max,
            "duckdb_path": str(duck_path),
        }
        db.commit()
        logger.info("CRISP dataset registered: %s (%d rows)", document.filename, row_count)
        return profile

    except Exception as exc:
        logger.warning("CRISP dataset registration failed for %s: %s", document.id, exc)
        document.analyst_profile = None
        db.commit()
        return None
    finally:
        con.close()


def _xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> None:
    from openpyxl import load_workbook
    import csv

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()

    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(["" if c is None else c for c in row])


def load_profile(document: Document) -> DatasetProfile | None:
    raw = document.analyst_profile
    if not raw:
        return None
    return DatasetProfile(
        document_id=document.id,
        document_name=document.filename,
        row_count=int(raw.get("row_count") or 0),
        columns=list(raw.get("columns") or []),
        date_column=raw.get("date_column"),
        quantity_column=raw.get("quantity_column"),
        price_column=raw.get("price_column"),
        category_column=raw.get("category_column"),
        city_column=raw.get("city_column"),
        product_column=raw.get("product_column"),
        date_min=raw.get("date_min"),
        date_max=raw.get("date_max"),
    )


def list_ready_datasets(db: Session, category: str | None = None) -> list[tuple[Document, DatasetProfile]]:
    q = (
        db.query(Document)
        .filter(Document.status == "ready", Document.analyst_profile.isnot(None))
    )
    if category and category.strip().lower() not in ("", "general"):
        q = q.filter(Document.category == category.strip().lower())

    out: list[tuple[Document, DatasetProfile]] = []
    for doc in q.all():
        profile = load_profile(doc)
        if profile and profile.row_count > 0:
            out.append((doc, profile))
    return out


def connect_dataset(document_id: uuid.UUID) -> duckdb.DuckDBPyConnection:
    path = _duckdb_path(document_id)
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found: {document_id}")
    return duckdb.connect(str(path))
