"""Document parsing and chunking."""

from __future__ import annotations

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

CHUNK_SIZE = 800
CHUNK_OVERLAP = 100


def _split_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []

    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start = end - overlap
        if start >= len(text):
            break
    return chunks


def parse_pdf(file_path: str) -> list[dict]:
    import fitz

    chunks_with_meta: list[dict] = []
    doc = fitz.open(file_path)
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text()
        for chunk_text in _split_text(text):
            chunks_with_meta.append({"content": chunk_text, "page": page_num + 1})
    doc.close()
    return chunks_with_meta


def parse_docx(file_path: str) -> list[dict]:
    from docx import Document as DocxDocument

    doc = DocxDocument(file_path)
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return [{"content": c, "page": None} for c in _split_text(full_text)]


def parse_csv(file_path: str) -> list[dict]:
    import csv

    lines: list[str] = []
    with open(file_path, encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return []

    header = [c.strip() for c in rows[0]]
    for row in rows[1:]:
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        if header and any(header):
            pairs = [f"{h}: {v}" for h, v in zip(header, cells) if h and v]
            lines.append(" | ".join(pairs) if pairs else " | ".join(cells))
        else:
            lines.append(" | ".join(cells))

    full_text = "\n".join(lines)
    # Una fila = un chunk para que vendedor/notas no se mezclen al buscar o graficar.
    return [{"content": line, "page": None} for line in lines if line.strip()]


def _sheet_rows_to_lines(rows: list[tuple]) -> list[str]:
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    lines: list[str] = []
    for row in rows[1:]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        if header and any(header):
            pairs = [f"{h}: {v}" for h, v in zip(header, cells) if h and v]
            lines.append(" | ".join(pairs) if pairs else " | ".join(cells))
        else:
            lines.append(" | ".join(cells))
    return lines


def parse_xlsx(file_path: str) -> list[dict]:
    from openpyxl import load_workbook

    chunks_with_meta: list[dict] = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            lines = _sheet_rows_to_lines(rows)
            if not lines:
                continue
            sheet_text = f"Hoja: {sheet.title}\n" + "\n".join(lines)
            for chunk_text in _split_text(sheet_text):
                chunks_with_meta.append({"content": chunk_text, "sheet": sheet.title})
    finally:
        wb.close()
    return chunks_with_meta


def parse_txt(file_path: str) -> list[dict]:
    text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
    return [{"content": c, "page": None} for c in _split_text(text)]


def parse_document(file_path: str, mime_type: str) -> list[dict]:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if mime_type == "application/pdf" or suffix == ".pdf":
        return parse_pdf(file_path)
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or suffix in (".docx", ".doc"):
        return parse_docx(file_path)
    if mime_type.startswith("text/") or suffix in (".txt", ".md"):
        return parse_txt(file_path)
    if suffix == ".csv" or mime_type == "text/csv":
        return parse_csv(file_path)
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or suffix in (".xlsx", ".xlsm"):
        return parse_xlsx(file_path)

    logger.warning("Unsupported mime type %s, attempting text read", mime_type)
    return parse_txt(file_path)
